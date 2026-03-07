import io
import requests
import pandas as pd
import time
import os
import shutil
import random
import warnings
import json
import sqlite3
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

# Google API 相關庫
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload

warnings.filterwarnings("ignore", message="Unverified HTTPS request")

# === 1. 權限與路徑設定 ===
# 這裡會從 GitHub Secrets 讀取你存入的 GCP_SERVICE_ACCOUNT_KEY
SERVICE_ACCOUNT_INFO = json.loads(os.environ.get('GCP_SERVICE_ACCOUNT_KEY'))
SCOPES = ['https://www.googleapis.com/auth/drive']
creds = service_account.Credentials.from_service_account_info(SERVICE_ACCOUNT_INFO, scopes=SCOPES)
drive_service = build('drive', 'v3', credentials=creds)

# 雲端資料夾 ID
BASE_DIR_ID = "1WE_tmOMxh9zffOVau4lAFOTnWDWFy_wJ"
BACKUP_DIR_ID = "11YNuJdVqUkAJTDhN4ASIkrxGDK3_PDUO"

# GitHub Actions 虛擬機內的臨時資料夾
LOCAL_TEMP = "./temp_stock"
DB_FILE_NAME = "all_stocks_data.db"
LOCAL_DB_PATH = os.path.join(LOCAL_TEMP, DB_FILE_NAME)
if not os.path.exists(LOCAL_TEMP):
    os.makedirs(LOCAL_TEMP)

session = requests.Session()
session.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
})

# === 2. 雲端 DB 同步工具 ===
def sync_db_from_cloud():
    """從雲端下載資料庫檔案到 GitHub 虛擬機"""
    query = f"name = '{DB_FILE_NAME}' and '{BASE_DIR_ID}' in parents and trashed = false"
    results = drive_service.files().list(q=query, fields="files(id)").execute()
    items = results.get('files', [])
    if items:
        file_id = items[0]['id']
        print(f"📥 正在從雲端下載資料庫 (ID: {file_id})...")
        request = drive_service.files().get_media(fileId=file_id)
        with io.FileIO(LOCAL_DB_PATH, 'wb') as fh:
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while not done:
                _, done = downloader.next_chunk()
        return file_id
    print("ℹ️ 雲端尚無資料庫檔案，稍後將建立新檔。")
    return None

def sync_db_to_cloud(existing_id=None):
    """將本地更新後的資料庫上傳覆蓋雲端"""
    print(f"📤 正在將資料庫同步至雲端...")
    media = MediaFileUpload(LOCAL_DB_PATH, mimetype='application/x-sqlite3')
    if existing_id:
        drive_service.files().update(fileId=existing_id, media_body=media).execute()
    else:
        file_metadata = {'name': DB_FILE_NAME, 'parents': [BASE_DIR_ID]}
        drive_service.files().create(body=file_metadata, media_body=media).execute()
    print("✅ 資料庫雲端同步完成！")
    
# === 3. 股票抓取功能 ===
def get_all_taiwan_stocks():
    def get_stocks(mode):
        url = f"https://isin.twse.com.tw/isin/C_public.jsp?strMode={mode}"
        res = session.get(url, verify=False)
        res.encoding = "big5"
        df = pd.read_html(io.StringIO(res.text))[0]
        df.columns = ["有價證券代號及名稱", "ISIN", "日期", "市場別", "產業別", "CFICode", "備註"]
        df = df[df["CFICode"] == "ESVUFR"]
        return {row.split()[0]: row.split()[1] for row in df["有價證券代號及名稱"] if len(row.split()) >= 2}
    
    all_s = {**get_stocks(2), **get_stocks(4)}
    print(f"📚 合計股票: {len(all_s)} 檔")
    return all_s

def fetch_price_by_volume(symbol="2330.TW"):
    """取得 Yahoo 股價價量分佈資料（自動容錯解析）"""
    url = f"https://tw.stock.yahoo.com/_td-stock/api/resource/StockServices.priceByVolumes;symbol={symbol}?bkt=%5B%22tw-stock-desktop-future-rampup%22%5D&device=desktop&lang=zh-Hant-TW&region=TW"
    headers = {"User-Agent": "Mozilla/5.0"}
    res = session.get(url, headers=headers)
    res.raise_for_status()
    js = res.json()

    # 判斷 pagination 裡面的 resultsTotal 是否為 0
    pagination = js.get('data', {}).get('pagination', {}) or js.get('pagination', {})
    if pagination.get('resultsTotal') == 0:
        print(f"ℹ️ {symbol} Yahoo 查無價量分佈資料 (resultsTotal: 0)")
        return {"date": None, "data": []} # 回傳空的 data

    # --------------------------
    #  1️⃣ 找 priceByVolumes
    # --------------------------
    data = None
    for path in [
        ["data", "priceByVolumes"],
        ["priceByVolumes"],
        ["data", "data", "priceByVolumes"],
    ]:
        d = js
        try:
            for k in path:
                d = d[k]
            if isinstance(d, list):
                data = d
                break
        except (KeyError, TypeError):
            continue

    if data == []:
        # 檢查 pagination 確認是否真的沒成交
        pagination = js.get('data', {}).get('pagination', {}) or js.get('pagination', {})
        if pagination.get('resultsTotal') == 0:
            print(f"ℹ️ {symbol} 正常回傳：今日無成交價量分佈 (resultsTotal: 0)")
            return {"date": None, "data": []}
            
    if not data:
        print("⚠️ 無法從 Yahoo 回傳中找到 priceByVolumes，實際回傳內容：")
        print(js)
        raise ValueError("Yahoo API 結構已更改，請檢查回傳格式。")

    price_list = [(float(d["price"]), int(d["volumeK"])) for d in data if d.get("volumeK")]

    # --------------------------
    #  2️⃣ 找 日期（多層容錯）
    # --------------------------
    date = None
    date_paths = [
        ["data", "date"],
        ["date"],
        ["data", "data", "date"],
        ["meta", "date"],
    ]

    for path in date_paths:
        d = js
        try:
            for k in path:
                d = d[k]
            if isinstance(d, str) and len(d) >= 10:
                date = d[:10]
                break
        except Exception:
            continue

    # 若仍然找不到，用今日系統日期（最後保底方案）
    if date is None:
        from datetime import datetime
        date = datetime.now().strftime("%Y-%m-%d")

    return {
        "date": date,
        "data": price_list,
    }


def fetch_margin_data(symbol="2330.TW"):
    """取得 Yahoo 融資融券資料（自動容錯解析）"""
    url = f"https://tw.stock.yahoo.com/_td-stock/api/resource/StockServices.creditsWithQuoteStats;limit=90;symbol={symbol}?bkt=%5B%22tw-stock-desktop-future-rampup%22%5D&device=desktop&lang=zh-Hant-TW&region=TW"
    headers = {"User-Agent": "Mozilla/5.0"}
    res = session.get(url, headers=headers)
    res.raise_for_status()
    js = res.json()

    # --- 多層防呆找資料 ---
    possible_paths = [
        ["data", "data", "result", "credits"],
        ["data", "result", "credits"],
        ["result", "credits"],
        ["credits"],
    ]

    credits = None
    for path in possible_paths:
        d = js
        try:
            for k in path:
                d = d[k]
            if isinstance(d, list) and len(d) > 0:
                credits = d[0]
                break
        except (KeyError, TypeError):
            continue

    if not credits:
        print("⚠️ 無法從 Yahoo 回傳中找到 credits 資料，實際回傳內容：")
        print(js)
        raise ValueError("Yahoo API 結構已更改，請檢查回傳格式。")

    # --- 萃取日期、融資、融券 ---
    date = credits.get("date", "").split("T")[0]
    fin = int(credits.get("financingTotalVolK", 0))
    short = int(credits.get("shortTotalVolK", 0))

    return {
        "date": date,
        "融資": fin,
        "融券": short,
    }

def safe_fetch_margin_data(symbol, prev_margin_data=None):
    """
    嘗試安全抓取 margin data；若失敗，回傳預設值。
    prev_margin_data: 昨日資料（若要計算融資差與融券差）
    """
    def safe_int(value):
        if value is None:
            return 0
        try:
            return int(value)
        except (ValueError, TypeError):
            return 0
            
    try:
        margin = fetch_margin_data(symbol)

        # 若有前一日資料，計算變化量
        if prev_margin_data:
            prev_fin = safe_int(prev_margin_data.get("融資"))
            prev_short = safe_int(prev_margin_data.get("融券"))
            margin["融資差"] = margin["融資"] - prev_fin
            margin["融券差"] = margin["融券"] - prev_short
        else:
            margin["融資差"] = 0
            margin["融券差"] = 0

        return margin

    except Exception as e:
        print(f"⚠️ [警告] 無法取得 {symbol} 的資券資料: {e}")
        # 若發生錯誤，回傳預設為 0 的結構
        return {
            "date": pd.Timestamp.now().strftime("%Y-%m-%d"),
            "融資": 0,
            "融資差": 0,
            "融券": 0,
            "融券差": 0,
        }

def _safe_raw(x):
    """將 {'raw': '21.5'} 或 '21.5' 或 None -> float 或 None"""
    try:
        if x is None:
            return None
        if isinstance(x, dict) and "raw" in x:
            return float(x["raw"])
        return float(x)
    except:
        return None
        
def fetch_ohlc_data(symbol="2330.TW"):
    """用 path 方式從 Yahoo StockServices.stockList 抓 OHLC"""

    url = (
        "https://tw.stock.yahoo.com/_td-stock/api/resource/StockServices.stockList"
        f";fields=avgPrice%2Corderbook;symbols={symbol}"
        "?device=desktop&ecma=modern&intl=tw&lang=zh-Hant-TW&returnMeta=true"
    )
    headers = {"User-Agent": "Mozilla/5.0"}
    res = requests.get(url, headers=headers, timeout=10)
    res.raise_for_status()
    js = res.json()

    # 可能的結構 path（Yahoo 偶爾會改結構）
    paths = [
        ["data", "list", 0],      # 最常見
        ["data", 0],              # 偶爾 Yahoo 改成這樣
        ["list", 0],              # 另外種可能
        ["data"],                 # 若只有一個物件直接 data
    ]

    item = None
    for path in paths:
        d = js
        try:
            for k in path:
                if isinstance(k, int):
                    d = d[k]
                else:
                    d = d[k]
            # 如果有 price 就當成成功
            if "price" in d:
                item = d
                break
        except Exception:
            continue

    if item is None:
        print("⚠️ Yahoo API 結構變更，找不到 OHLC 資料。回傳內容：")
        print(js)
        raise ValueError("無法從 Yahoo 回傳取得 OHLC 欄位")

    # --- OHLC ---
    pre_close_val = _safe_raw(item.get("regularMarketPreviousClose"))
    high_val  = _safe_raw(item.get("regularMarketDayHigh"))
    low_val  = _safe_raw(item.get("regularMarketDayLow"))
    open_val   = _safe_raw(item.get("regularMarketOpen"))
    close_val = _safe_raw(item.get("price"))

    # --- regularMarketTime 取日期 ---
    time_str = item.get("regularMarketTime")
    if isinstance(time_str, str) and len(time_str) >= 10:
        date = time_str[:10]
    else:
        date = None  # or today's date if needed

    return {
        "date": date,
        "開": open_val,
        "高": high_val,
        "低": low_val,
        "收": close_val,
        "昨收": pre_close_val
    }

# === 4. Excel 處理邏輯 (雲端同步版) ===
def update_excel_and_db_cloud(symbol, name, price_data, margin_data, ohlc_data):
    file_name = f"{symbol}_價量報表.xlsx"
    local_path = os.path.join(LOCAL_TEMP, file_name)
    today = margin_data["date"]
    
    # A. 搜尋雲端是否有舊檔
    query = f"name = '{file_name}' and '{BASE_DIR_ID}' in parents and trashed = false"
    results = drive_service.files().list(q=query, fields="files(id)").execute()
    items = results.get('files', [])

    # B. 下載或建立
    if items:
        file_id = items[0]['id']
        request = drive_service.files().get_media(fileId=file_id)
        with io.FileIO(local_path, 'wb') as fh:
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while not done:
                _, done = downloader.next_chunk()
        wb = load_workbook(local_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "價量報表"
        ws["A1"], ws["B1"] = symbol, name
        ws["A2"], ws["A3"], ws["A4"], ws["A5"] = "融資", "融資差", "融券", "融券差"
        ws["A7"], ws["A8"], ws["A9"], ws["A10"] = "開", "高", "低", "收"
        ws["A12"], ws["B12"] = "成交價", "歷史成交量總和"
        file_id = None

    # C. 寫入與著色邏輯
    headers = [ws.cell(row=1, column=c).value for c in range(3, ws.max_column + 1)]
    if today not in headers:
        ws.cell(row=1, column=ws.max_column + 1).value = today
    today_col = headers.index(today) + 3 if today in headers else ws.max_column

    prev_col = today_col - 1 if today_col > 3 else None
    prev_fin = ws.cell(row=2, column=prev_col).value if prev_col else 0
    prev_short = ws.cell(row=4, column=prev_col).value if prev_col else 0

    ws.cell(row=2, column=today_col, value=margin_data["融資"])
    ws.cell(row=3, column=today_col, value=margin_data["融資"] - (prev_fin or 0))
    ws.cell(row=4, column=today_col, value=margin_data["融券"])
    ws.cell(row=5, column=today_col, value=margin_data["融券"] - (prev_short or 0))

    ws.cell(row=7, column=today_col, value=ohlc_data["開"])
    ws.cell(row=8, column=today_col, value=ohlc_data["高"])
    ws.cell(row=9, column=today_col, value=ohlc_data["低"])
    ws.cell(row=10, column=today_col, value=ohlc_data["收"])

    # 價量更新
    existing_prices = {}
    for r in range(13, ws.max_row + 1):
        price = ws.cell(row=r, column=1).value
        if price:
            existing_prices[float(price)] = r

    today_prices = [p for p, _ in price_data]
    all_prices = sorted(set(existing_prices.keys()) | set(today_prices), reverse=True)

    current_row = 13
    new_map = {}
    for price in all_prices:
        if price not in existing_prices:
            ws.insert_rows(current_row)
            ws.cell(row=current_row, column=1, value=price)
            ws.cell(row=current_row, column=2, value=0)
        new_map[price] = current_row
        current_row += 1

    for price, vol in price_data:
        ws.cell(row=new_map[price], column=today_col, value=vol)

    for price in all_prices:
        row = new_map[price]
        total = 0
        for c in range(3, ws.max_column + 1):
            val = ws.cell(row=row, column=c).value
            total += val if isinstance(val, (int, float)) else 0
        ws.cell(row=row, column=2, value=total)

    # D. 著色邏輯 (完全同步 Local 版)
    fill_row_top_1_5 = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    fill_row_top_6_10 = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    fill_row_top_11_20 = PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid")
    fill_cell_top_1_5 = PatternFill(start_color="FF7F50", end_color="FF7F50", fill_type="solid")
    fill_cell_top_6_10 = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    fill_close_price = PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid")
    fill_open_price = PatternFill(start_color="FF00FF", end_color="FF00FF", fill_type="solid")
    fill_none = PatternFill(fill_type=None)

    start_data_row = 13
    end_data_row = ws.max_row
    avg_volume_list = []
    raw_volume_list = []

    for r in range(start_data_row, end_data_row + 1):
        # 清除舊色
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).fill = fill_none
        
        raw_val = ws.cell(row=r, column=2).value or 0
        raw_volume_list.append((r, raw_val))

        window_start = max(start_data_row, r - 2)
        window_end = min(end_data_row, r + 2)
        window_values = [ws.cell(row=wr, column=2).value or 0 for wr in range(window_start, window_end + 1)]
        avg_val = sum(window_values) / len(window_values) if window_values else 0
        avg_volume_list.append((r, avg_val))

    # 第一階段：平均值上色 (整列)
    avg_volume_list.sort(key=lambda x: x[1], reverse=True)
    for i in range(min(20, len(avg_volume_list))):
        target_row = avg_volume_list[i][0]
        if i < 5: fill = fill_row_top_1_5
        elif i < 10: fill = fill_row_top_6_10
        else: fill = fill_row_top_11_20
        for c in range(1, ws.max_column + 1):
            ws.cell(row=target_row, column=c).fill = fill

    # 第二階段：原始值上色 (B欄強化)
    raw_volume_list.sort(key=lambda x: x[1], reverse=True)
    for i in range(min(10, len(raw_volume_list))):
        target_row = raw_volume_list[i][0]
        ws.cell(row=target_row, column=2).fill = fill_cell_top_1_5 if i < 5 else fill_cell_top_6_10

    # 第三階段：開收盤價 (A欄)
    close_price = float(ohlc_data["收"])
    open_price = float(ohlc_data["開"])
    if open_price in new_map:
        ws.cell(row=new_map[open_price], column=1).fill = fill_open_price
    if close_price in new_map:
        ws.cell(row=new_map[close_price], column=1).fill = fill_close_price

    wb.save(local_path)

    # D. 上傳回雲端
    media = MediaFileUpload(local_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if file_id:
        drive_service.files().update(fileId=file_id, media_body=media).execute()
    else:
        file_metadata = {'name': file_name, 'parents': [BASE_DIR_ID]}
        drive_service.files().create(body=file_metadata, media_body=media).execute()
    print(f"✅ 已同步雲端: {file_name}")
    # --- Part B: SQLite 資料庫維護 ---
    conn = sqlite3.connect(LOCAL_DB_PATH)
    cursor = conn.cursor()
    try:
        # 1. 每日信用交易
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS daily_credit_trading (
                stock_id TEXT, trade_date TEXT, 
                margin_balance INTEGER, margin_change INTEGER, 
                short_balance INTEGER, short_change INTEGER,
                PRIMARY KEY (stock_id, trade_date)
            )
        """)
        cursor.execute("INSERT OR REPLACE INTO daily_credit_trading VALUES (?,?,?,?,?,?)", (
            symbol, today, margin_data["融資"], margin_data.get("融資差", 0), 
            margin_data["融券"], margin_data.get("融券差", 0)
        ))

        # 2. 每日開高低收
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS daily_stock_prices (
                stock_id TEXT, trade_date TEXT, 
                open_price REAL, high_price REAL, low_price REAL, close_price REAL,
                PRIMARY KEY (stock_id, trade_date)
            )
        """)
        cursor.execute("INSERT OR REPLACE INTO daily_stock_prices VALUES (?,?,?,?,?,?)", (
            symbol, today, ohlc_data["開"], ohlc_data["高"], ohlc_data["低"], ohlc_data["收"]
        ))

        # 3. 每日價格成交分佈
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS daily_price_volume_distribution (
                stock_id TEXT, trade_date TEXT, price REAL, volume INTEGER,
                PRIMARY KEY (stock_id, trade_date, price)
            )
        """)
        for p, v in price_data:
            cursor.execute("INSERT OR REPLACE INTO daily_price_volume_distribution VALUES (?,?,?,?)", 
                           (symbol, today, p, v))
        conn.commit()
    except Exception as e:
        print(f"❌ DB 更新失敗 {symbol}: {e}")
    finally:
        conn.close()
    print(f"✅ 已同步雲端 Excel: {symbol}")

# === 5. 主程式 ===
def main():
    # 步驟 1: 下載雲端現有資料庫
    db_cloud_id = sync_db_from_cloud()
    stocks = get_all_taiwan_stocks()
    for symbol, name in list(stocks.items()):
#    for symbol, name in list(stocks.items())[:5]:  # 測試先跑前5檔
        retry_count = 0
        max_retries = 24  # 最多重試 3 次
        success = False
            
        while retry_count <= max_retries:
            try:
                print(f"\n📈 處理 {symbol} {name} (嘗試次數: {retry_count + 1})")
                p = fetch_price_by_volume(symbol)
                if not p['data']: # 如果是 []
                    break
                m = safe_fetch_margin_data(symbol)
                o = fetch_ohlc_data(symbol)
    
                # 比對日期
                if p['date'] == m['date'] == o['date']:
                    update_excel_and_db_cloud(symbol, name, p["data"], m, o)
                    time.sleep(random.uniform(1.0, 2.0))
                    success = True
                    break
                else:
                    #print(f"⚠️ {symbol} 日期不一致，跳過更新 (p:{p['date']}, m:{m['date']}, o:{o['date']})")
                    if symbol == "1101"  #第一檔，1101台泥，必定有成交量
                        retry_count += 1
                        if retry_count <= max_retries:
                            print(f"⏳ {symbol} 日期不一致 (p:{p['date']}, m:{m['date']}, o:{o['date']})，"
                                  f"第 {retry_count} 次重試，等待 300 秒...")
                            time.sleep(300) # 等待 Yahoo 更新
                        else:
                            print(f"❌ {symbol} 達到重試上限，跳過更新。")
                            break
                    else
                        print(f"⏳ {symbol} 日期不一致 (p:{p['date']}, m:{m['date']}, o:{o['date']})")
            except (json.JSONDecodeError, ValueError, requests.exceptions.RequestException) as e:
                    # --- 被擋 IP 或網路錯誤情境 (你圖中 line 1 column 1 的錯誤) ---
                    if retry_count < max_retries:
                        # 被擋時採取「指數型等待」，第一次 60s, 第二次 120s...
                        #wait_time = (retry_count + 5) * 60 + random.randint(1, 15)
                        wait_time = 300
                        print(f"⚠️ {symbol} 抓取遭拒或格式錯誤: {e}")
                        print(f"🛑 觸發冷卻機制，等待 {wait_time} 秒後重試...")
                        time.sleep(wait_time)
                        retry_count += 1
                    else:
                        print(f"❌ {symbol} 連續失敗，放棄處理。")
                        break
            except Exception as e:
                print(f"❌ {symbol} 失敗: {e}")
                break

    # 步驟 2: 建立索引並上傳 DB
    if os.path.exists(LOCAL_DB_PATH):
        conn = sqlite3.connect(LOCAL_DB_PATH)
        cursor = conn.cursor()
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_credit ON daily_credit_trading (stock_id, trade_date);")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_prices ON daily_stock_prices (stock_id, trade_date);")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_dist ON daily_price_volume_distribution (stock_id, trade_date, price);")
        conn.close()
        sync_db_to_cloud(db_cloud_id)
        print(f"✅ 已同步雲端 DB")
if __name__ == "__main__":
    main()
